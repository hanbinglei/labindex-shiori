"""
LabIndex Shiori — XLSX 解析器

提取所有单元格文本（供全文检索）。
不强求标题/关键词，仅提取纯文本内容。
只读保障: 仅 open(path, 'rb')
"""

from __future__ import annotations

import logging
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)


def parse_xlsx(file_path: str) -> Dict:
    """
    解析 XLSX 文件，提取所有单元格文本

    Returns:
        dict: title=None, keywords=None, abstract=所有单元格文本,
              year=None, authors=None, note=None
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)

    all_text = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    text = str(cell.value).strip()
                    if text:
                        all_text.append(text)

    wb.close()

    combined = " ".join(all_text) if all_text else None

    return {
        "title": None,
        "title_source": None,
        "keywords": None,
        "keywords_source": None,
        "abstract": combined[:5000] if combined else None,  # 截断到5000字
        "year": None,
        "authors": None,
        "references": None,
        "note": None,
    }
